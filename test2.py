from tkinter import *
from tkinter import ttk
from openpyxl import *
import pandas as pd
from tkinter import messagebox
from fpdf import FPDF
from PIL import Image, ImageTk
import glob


class Student:
    def __init__(self, root):
        self.root = root
        self.root.title("Student Management System")
        width = 1350
        height = 820
        screen_width = root.winfo_screenwidth()
        screen_height = root.winfo_screenheight()
        x = (screen_width/2) - (width/2)
        y = (screen_height/2) - (height/2)
        self.root.geometry("%dx%d+%d+%d" % (width, height, x, y))
        self.root.resizable(False, False)

        title = Label(self.root, text="UNIQUE ENGLISH SCHOOL", bd=10, relief=GROOVE, font=(
            "Helvetica", 40, "bold"), bg="RoyalBlue3", fg="gold")
        title.pack(side=TOP, fill=X)

        load = Image.open("logo.png")
        load = load.resize((50, 45), Image.ANTIALIAS)
        render = ImageTk.PhotoImage(load)
        img = Label(self.root, image=render)
        img.image = render
        img.place(x=150, y=18)

        # ====================ALL VARIABLES========================

        self.roll_no_var = StringVar()
        self.name_var = StringVar()
        self.class_var = StringVar()
        self.section_var = StringVar()
        self.fname_var = StringVar()
        self.mname_var = StringVar()
        self.contact_var = StringVar()
        self.address_var = StringVar()
        self.branch_var = StringVar()
        self.house_var = StringVar()
        self.dob_var = StringVar()
        self.aadhar_var = StringVar()
        self.caste_var = StringVar()
        self.category_var = StringVar()
        self.bpl_var = StringVar()

        self.search_name = StringVar()
        self.search_class = StringVar()

        # =======================All Functions=====================

        wb = load_workbook('ues.xlsx')

        sheet = wb.active

        def excel():
            sheet.cell(row=1, column=1).value = "Roll"
            sheet.cell(row=1, column=2).value = "Name"
            sheet.cell(row=1, column=3).value = "Class"
            sheet.cell(row=1, column=4).value = "Section"
            sheet.cell(row=1, column=5).value = "F.Name"
            sheet.cell(row=1, column=16).value = "M.Name"
            sheet.cell(row=1, column=7).value = "Contact"
            sheet.cell(row=1, column=8).value = "Address"
            sheet.cell(row=1, column=9).value = "Branch"
            sheet.cell(row=1, column=10).value = "House"
            sheet.cell(row=1, column=11).value = "D.O.B"
            sheet.cell(row=1, column=12).value = "Aadhar"
            sheet.cell(row=1, column=13).value = "Caste"
            sheet.cell(row=1, column=14).value = "Category"
            sheet.cell(row=1, column=15).value = "BPL"

        def focus1(event):
            txt_name.focus_set()

        def focus2(event):
            txt_class.focus_set()

        def focus3(event):
            txt_Roll.focus_set()

        def focus4(event):
            txt_sec.focus_set()

        def focus5(event):
            txt_fname.focus_set()

        def focus6(event):
            txt_mname.focus_set()

        def focus7(event):
            txt_contact.focus_set()

        def focus8(event):
            txt_address.focus_set()

        def focus9(event):
            txt_branch.focus_set()

        def focus10(event):
            txt_house.focus_set()

        def focus11(event):
            txt_dob.focus_set()

        def focus12(event):
            txt_aadhar.focus_set()

        def focus13(event):
            txt_caste.focus_set()

        def focus14(event):
            txt_category.focus_set()

        def focus15(event):
            txt_bpl.focus_set()

        def clear():
            txt_name.delete(0, END)
            txt_class.delete(0, END)
            txt_Roll.delete(0, END)
            txt_sec.delete(0, END)
            txt_fname.delete(0, END)
            txt_mname.delete(0, END)
            txt_contact.delete(0, END)
            txt_address.delete(0, END)
            txt_branch.delete(0, END)
            txt_house.delete(0, END)
            txt_dob.delete(0, END)
            txt_aadhar.delete(0, END)
            txt_caste.delete(0, END)
            txt_category.delete(0, END)
            txt_bpl.delete(0, END)

        def Start():
            fp = pd.read_excel("./ues.xlsx")  # Read xlsx file
            # use for loop to get values in each line, _ is the number of line.
            if(Student_table):
                Student_table.delete(*Student_table.get_children())
                for _ in range(len(fp.index.values)):
                    Student_table.insert('', 'end', value=tuple(
                        fp.iloc[_, [0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14]].values))

        def insert_data():
            if(txt_Roll.get() == "" or
                    txt_name.get() == "" or
                    txt_class.get() == "" or
                    txt_sec.get() == "" or
                    txt_fname.get() == "" or
                    txt_mname.get() == "" or
                    txt_contact.get() == "" or
                    txt_address.get() == "" or
                    txt_branch.get() == "" or
                    txt_house.get() == "" or
                    txt_dob.get() == "" or
                    txt_aadhar.get() == "" or
                    txt_caste.get() == "" or
                    txt_category.get() == "" or
                    txt_bpl.get() == ""):

                messagebox.showerror("Error", "All Fields are Required!!!")

            else:
                current_row = sheet.max_row
                current_column = sheet.max_column

                sheet.cell(row=current_row + 1,
                           column=1).value = txt_Roll.get()
                sheet.cell(row=current_row + 1,
                           column=2).value = txt_name.get()
                sheet.cell(row=current_row + 1,
                           column=3).value = txt_class.get()
                sheet.cell(row=current_row + 1,
                           column=4).value = txt_sec.get()
                sheet.cell(row=current_row + 1,
                           column=5).value = txt_fname.get()
                sheet.cell(row=current_row + 1,
                           column=6).value = txt_mname.get()
                sheet.cell(row=current_row + 1,
                           column=7).value = txt_contact.get()
                sheet.cell(row=current_row + 1,
                           column=8).value = txt_address.get()
                sheet.cell(row=current_row + 1,
                           column=9).value = txt_branch.get()
                sheet.cell(row=current_row + 1,
                           column=10).value = txt_house.get()
                sheet.cell(row=current_row + 1,
                           column=11).value = txt_dob.get()
                sheet.cell(row=current_row + 1,
                           column=12).value = txt_aadhar.get()
                sheet.cell(row=current_row + 1,
                           column=13).value = txt_caste.get()
                sheet.cell(row=current_row + 1,
                           column=14).value = txt_category.get()
                sheet.cell(row=current_row + 1,
                           column=15).value = txt_bpl.get()

                # save file
                wb.save('ues.xlsx')

                txt_name.focus_set()

                Start()

                clear()

        def search_in_excel():
            df1 = pd.read_excel("./ues.xlsx")
            a1 = self.name_var.get()
            a2 = self.aadhar_var.get()

            for i, r in df1.iterrows():
                # print(r)
                if a1 in f"{r['Name']}" and a2 in f"{r['Aadhar']}":
                    self.row_num = i
                    return self.row_num

        def imgOpen():
            global st_img
            rollVAr = txt_Roll.get()
            classVar = txt_class.get()
            load1 = Image.open(f"Images/{rollVAr}{classVar}.jpg")
            load1 = load1.resize((89, 86), Image.ANTIALIAS)
            render = ImageTk.PhotoImage(load1)
            st_img = Label(Photo_Frame, image=render)
            st_img.image = render
            st_img.place(x=0, y=0)

        def fetch_image():
            rollVAr = txt_Roll.get()
            classVar = txt_class.get()

            path = "E:/Python Project/UES - Student Management System/Images"

            files = glob.glob(path + "/**/*.jpg", recursive=True)
            for file in files:
                if file.endswith(f'{rollVAr}{classVar}.jpg'):
                    print(str(file))
                    return imgOpen()

            # st_img.pack_forget()

            img_lbl = Label(Photo_Frame, text="Image Not Found",
                            bg="white", fg="black", font=("Helvetica", 7, "bold"))
            img_lbl.place(x=0, y=0, width=89, height=86)

        def get_cursor(ev):
            cursor_row = Student_table.focus()
            contents = Student_table.item(cursor_row)
            row = contents['values']

            self.roll_no_var.set(row[0])
            self.name_var.set(row[1])
            self.class_var.set(row[2])
            self.section_var.set(row[3])
            self.fname_var.set(row[4])
            self.mname_var.set(row[5])
            self.contact_var.set(row[6])
            self.address_var.set(row[7])
            self.branch_var.set(row[8])
            self.house_var.set(row[9])
            self.dob_var.set(row[10])
            self.aadhar_var.set(row[11])
            self.caste_var.set(row[12])
            self.category_var.set(row[13])
            self.bpl_var.set(row[14])

            fetch_image()

        def updateData():
            if(txt_Roll.get() == "" or
                    txt_name.get() == "" or
                    txt_class.get() == "" or
                    txt_sec.get() == "" or
                    txt_fname.get() == "" or
                    txt_mname.get() == "" or
                    txt_contact.get() == "" or
                    txt_address.get() == "" or
                    txt_branch.get() == "" or
                    txt_house.get() == "" or
                    txt_dob.get() == "" or
                    txt_aadhar.get() == "" or
                    txt_caste.get() == "" or
                    txt_category.get() == "" or
                    txt_bpl.get() == ""):

                messagebox.showerror("Error", "All Fields are Required!!!")

            else:
                search_in_excel()
                current_row = self.row_num + 2

                sheet.cell(row=current_row,
                           column=1).value = self.roll_no_var.get()
                sheet.cell(row=current_row,
                           column=2).value = self.name_var.get()
                sheet.cell(row=current_row,
                           column=3).value = self.class_var.get()
                sheet.cell(row=current_row,
                           column=4).value = self.section_var.get()
                sheet.cell(row=current_row,
                           column=5).value = self.fname_var.get()
                sheet.cell(row=current_row,
                           column=6).value = self.mname_var.get()
                sheet.cell(row=current_row,
                           column=7).value = self.contact_var.get()
                sheet.cell(row=current_row,
                           column=8).value = self.address_var.get()
                sheet.cell(row=current_row,
                           column=9).value = self.branch_var.get()
                sheet.cell(row=current_row,
                           column=10).value = self.house_var.get()
                sheet.cell(row=current_row,
                           column=11).value = self.dob_var.get()
                sheet.cell(row=current_row,
                           column=12).value = self.aadhar_var.get()
                sheet.cell(row=current_row,
                           column=13).value = self.caste_var.get()
                sheet.cell(row=current_row,
                           column=14).value = self.category_var.get()
                sheet.cell(row=current_row,
                           column=15).value = self.bpl_var.get()

                # save file
                wb.save('ues.xlsx')

                Start()

                clear()

        def deleteData():
            if(txt_Roll.get() == "" or
                    txt_name.get() == "" or
                    txt_class.get() == "" or
                    txt_sec.get() == "" or
                    txt_fname.get() == "" or
                    txt_mname.get() == "" or
                    txt_contact.get() == "" or
                    txt_address.get() == "" or
                    txt_branch.get() == "" or
                    txt_house.get() == "" or
                    txt_dob.get() == "" or
                    txt_aadhar.get() == "" or
                    txt_caste.get() == "" or
                    txt_category.get() == "" or
                    txt_bpl.get() == ""):

                messagebox.showerror(
                    "Error", "Data is not Found for Deletion!!")

            else:
                search_in_excel()
                current_row = self.row_num + 2

                sheet.delete_rows(current_row, 1)

                # save file
                wb.save('ues.xlsx')

                Start()

                clear()

        def searchData():
            query1 = txt_searchName.get()
            query2 = txt_searchClass.get()
            query3 = txt_searchRoll.get()
            selections = []
            for child in Student_table.get_children():
                if not query1.lower() in Student_table.item(child)['values'][1].lower() or not query2.lower() in str(Student_table.item(child)['values'][2]).lower() or not query3 in str(Student_table.item(child)['values'][0]):
                    selections.append(child)
                    Student_table.detach(child)

            Student_table.selection_set(selections)
            txt_searchName.delete(0, END)
            txt_searchClass.delete(0, END)
            txt_searchRoll.delete(0, END)

        # ======================MANAGE FRAME=======================

        excel()

        Photo_Frame = Frame(self.root, bd=4, relief=RIDGE, bg="white")
        Photo_Frame.place(x=150, y=83, width=90, height=87)

        Manage_Frame = Frame(self.root, bd=4, relief=RIDGE, bg="green4")
        Manage_Frame.place(x=10, y=170, width=390, height=640)

        m_title = Label(Manage_Frame, text="Manage Students", bg="green4", fg="white", font=(
            "Helvetica", 15, "bold"))
        m_title.grid(row=0, columnspan=2, pady=1)

        lbl_name = Label(Manage_Frame, text="Name", bg="green4", fg="white", font=(
            "Helvetica", 10, "bold"))
        lbl_name.grid(row=1, column=0, pady=5, padx=40, sticky="w")

        txt_name = Entry(Manage_Frame, textvariable=self.name_var, font=(
            "Helvetica", 10, "bold"), bd=4, relief=GROOVE)
        txt_name.grid(row=1, column=1, pady=5, padx=10, sticky="w")

        lbl_class = Label(Manage_Frame, text="Class", bg="green4", fg="white", font=(
            "Helvetica", 10, "bold"))
        lbl_class.grid(row=2, column=0, pady=5, padx=40, sticky="w")

        txt_class = Entry(Manage_Frame, textvariable=self.class_var, font=(
            "Helvetica", 10, "bold"), bd=4, relief=GROOVE)
        txt_class.grid(row=2, column=1, pady=5, padx=10, sticky="w")

        lbl_roll = Label(Manage_Frame, text="Roll No.", bg="green4", fg="white", font=(
            "Helvetica", 10, "bold"))
        lbl_roll.grid(row=3, column=0, pady=5, padx=40, sticky="w")

        txt_Roll = Entry(Manage_Frame, textvariable=self.roll_no_var, font=(
            "Helvetica", 10, "bold"), bd=4, relief=GROOVE)
        txt_Roll.grid(row=3, column=1, pady=5, padx=10, sticky="w")

        lbl_sec = Label(Manage_Frame, text="Section", bg="green4", fg="white", font=(
            "Helvetica", 10, "bold"))
        lbl_sec.grid(row=4, column=0, pady=5, padx=40, sticky="w")

        txt_sec = Entry(Manage_Frame, textvariable=self.section_var, font=(
            "Helvetica", 10, "bold"), bd=4, relief=GROOVE)
        txt_sec.grid(row=4, column=1, pady=5, padx=10, sticky="w")

        lbl_fname = Label(Manage_Frame, text="F. Name", bg="green4", fg="white", font=(
            "Helvetica", 10, "bold"))
        lbl_fname.grid(row=5, column=0, pady=5, padx=40, sticky="w")

        txt_fname = Entry(Manage_Frame, textvariable=self.fname_var, font=(
            "Helvetica", 10, "bold"), bd=4, relief=GROOVE)
        txt_fname.grid(row=5, column=1, pady=5, padx=10, sticky="w")

        lbl_mname = Label(Manage_Frame, text="M. Name", bg="green4", fg="white", font=(
            "Helvetica", 10, "bold"))
        lbl_mname.grid(row=6, column=0, pady=5, padx=40, sticky="w")

        txt_mname = Entry(Manage_Frame, textvariable=self.mname_var, font=(
            "Helvetica", 10, "bold"), bd=4, relief=GROOVE)
        txt_mname.grid(row=6, column=1, pady=5, padx=10, sticky="w")

        lbl_contact = Label(Manage_Frame, text="Contact", bg="green4", fg="white", font=(
            "Helvetica", 10, "bold"))
        lbl_contact.grid(row=7, column=0, pady=5, padx=40, sticky="w")

        txt_contact = Entry(Manage_Frame, textvariable=self.contact_var, font=(
            "Helvetica", 10, "bold"), bd=4, relief=GROOVE)
        txt_contact.grid(row=7, column=1, pady=5, padx=10, sticky="w")

        lbl_address = Label(Manage_Frame, text="Address", bg="green4", fg="white", font=(
            "Helvetica", 10, "bold"))
        lbl_address.grid(row=8, column=0, pady=5, padx=40, sticky="w")

        txt_address = Entry(Manage_Frame, textvariable=self.address_var, width=20, font=(
            "Helvetica", 10, "bold"), bd=4, relief=GROOVE)
        txt_address.grid(row=8, column=1, pady=5, padx=10, sticky="w")

        lbl_branch = Label(Manage_Frame, text="Branch", bg="green4", fg="white", font=(
            "Helvetica", 10, "bold"))
        lbl_branch.grid(row=9, column=0, pady=5, padx=40, sticky="w")

        txt_branch = Entry(Manage_Frame, textvariable=self.branch_var, font=(
            "Helvetica", 10, "bold"), bd=4, relief=GROOVE)
        txt_branch.grid(row=9, column=1, pady=5, padx=10, sticky="w")

        lbl_house = Label(Manage_Frame, text="House", bg="green4", fg="white", font=(
            "Helvetica", 10, "bold"))
        lbl_house.grid(row=10, column=0, pady=5, padx=40, sticky="w")

        txt_house = Entry(Manage_Frame, textvariable=self.house_var, font=(
            "Helvetica", 10, "bold"), bd=4, relief=GROOVE)
        txt_house.grid(row=10, column=1, pady=5, padx=10, sticky="w")

        lbl_dob = Label(Manage_Frame, text="D.O.B", bg="green4", fg="white", font=(
            "Helvetica", 10, "bold"))
        lbl_dob.grid(row=11, column=0, pady=5, padx=40, sticky="w")

        txt_dob = Entry(Manage_Frame, textvariable=self.dob_var, font=(
            "Helvetica", 10, "bold"), bd=4, relief=GROOVE)
        txt_dob.grid(row=11, column=1, pady=5, padx=10, sticky="w")

        lbl_aadhar = Label(Manage_Frame, text="Aadhar No.", bg="green4", fg="white", font=(
            "Helvetica", 10, "bold"))
        lbl_aadhar.grid(row=12, column=0, pady=5, padx=40, sticky="w")

        txt_aadhar = Entry(Manage_Frame, textvariable=self.aadhar_var, font=(
            "Helvetica", 10, "bold"), bd=4, relief=GROOVE)
        txt_aadhar.grid(row=12, column=1, pady=5, padx=10, sticky="w")

        lbl_caste = Label(Manage_Frame, text="Caste", bg="green4", fg="white", font=(
            "Helvetica", 10, "bold"))
        lbl_caste.grid(row=13, column=0, pady=5, padx=40, sticky="w")

        txt_caste = Entry(Manage_Frame, textvariable=self.caste_var, font=(
            "Helvetica", 10, "bold"), bd=4, relief=GROOVE)
        txt_caste.grid(row=13, column=1, pady=5, padx=10, sticky="w")

        lbl_category = Label(Manage_Frame, text="Category", bg="green4", fg="white", font=(
            "Helvetica", 10, "bold"))
        lbl_category.grid(row=14, column=0, pady=5, padx=40, sticky="w")

        txt_category = Entry(Manage_Frame, textvariable=self.category_var, font=(
            "Helvetica", 10, "bold"), bd=4, relief=GROOVE)
        txt_category.grid(row=14, column=1, pady=5, padx=10, sticky="w")

        lbl_bpl = Label(Manage_Frame, text="BPL", bg="green4", fg="white", font=(
            "Helvetica", 10, "bold"))
        lbl_bpl.grid(row=15, column=0, pady=5, padx=40, sticky="w")

        txt_bpl = Entry(Manage_Frame, textvariable=self.bpl_var, font=(
            "Helvetica", 10, "bold"), bd=4, relief=GROOVE)
        txt_bpl.grid(row=15, column=1, pady=5, padx=10, sticky="w")

        txt_name.bind("<Return>", focus1)
        txt_class.bind("<Return>", focus2)
        txt_Roll.bind("<Return>", focus3)
        txt_sec.bind("<Return>", focus4)
        txt_fname.bind("<Return>", focus5)
        txt_mname.bind("<Return>", focus6)
        txt_contact.bind("<Return>", focus7)
        txt_address.bind("<Return>", focus8)
        txt_branch.bind("<Return>", focus9)
        txt_house.bind("<Return>", focus10)
        txt_dob.bind("<Return>", focus11)
        txt_aadhar.bind("<Return>", focus12)
        txt_caste.bind("<Return>", focus13)
        txt_category.bind("<Return>", focus14)
        txt_bpl.bind("<Return>", focus15)

        # ====================BUTTON FRAME=========================

        btn_frame = Frame(Manage_Frame, bd=4, relief=RIDGE, bg="green4")
        btn_frame.place(x=7, y=580, width=365)

        addBtn = Button(btn_frame, text="Add", width=7, bg="deep sky blue",
                        command=insert_data)
        addBtn.grid(
            row=0, column=0, padx=5, pady=7)
        updateBtn = Button(btn_frame, text="Update", bg="dark orange",
                           width=7, command=updateData)
        updateBtn.grid(
            row=0, column=1, padx=5, pady=7)
        deleteBtn = Button(btn_frame, text="Delete", bg="orange red",
                           width=7, command=deleteData)
        deleteBtn.grid(
            row=0, column=2, padx=5, pady=7)
        clearBtn = Button(btn_frame, text="Clear", width=7,
                          bg="PaleGreen1", command=clear)
        clearBtn.grid(
            row=0, column=3, padx=5, pady=7)

        # =======================DETAIL FRAME======================

        Detail_Frame = Frame(self.root, bd=4, relief=RIDGE, bg="green4")
        Detail_Frame.place(x=410, y=100, width=925, height=710)

        lbl_search = Label(Detail_Frame, text="Search :", bg="green4", fg="white", font={
                           "Helvetica", 15, "bold"})
        lbl_search.grid(row=0, column=0, padx=10, pady=10)

        lbl_searchName = Label(Detail_Frame, text="Name", bg="green4", fg="white", font=(
            "Helvetica", 12, "bold"))
        lbl_searchName.grid(row=0, column=1, pady=5, padx=5, sticky="w")

        txt_searchName = Entry(Detail_Frame, font=(
            "Helvetica", 10, "bold"), bd=5, relief=GROOVE)
        txt_searchName.grid(row=0, column=2, pady=5, padx=2, sticky="w")

        lbl_searchClass = Label(Detail_Frame, text="Class", bg="green4", fg="white", font=(
            "Helvetica", 12, "bold"))
        lbl_searchClass.grid(row=0, column=3, pady=5, padx=8, sticky="w")

        txt_searchClass = Entry(Detail_Frame, font=(
            "Helvetica", 10, "bold"), bd=5, relief=GROOVE)
        txt_searchClass.grid(row=0, column=4, pady=5, padx=2, sticky="w")

        lbl_searchRoll = Label(Detail_Frame, text="Roll", bg="green4", fg="white", font=(
            "Helvetica", 12, "bold"))
        lbl_searchRoll.grid(row=0, column=5, pady=5, padx=8, sticky="w")

        txt_searchRoll = Entry(Detail_Frame, font=(
            "Helvetica", 10, "bold"), bd=5, relief=GROOVE)
        txt_searchRoll.grid(row=0, column=6, pady=5, padx=2, sticky="w")

        searchBtn = Button(Detail_Frame, text="Search",
                           width=7, command=searchData)
        searchBtn.grid(
            row=0, column=7, padx=15, pady=12)
        refreshBtn = Button(Detail_Frame, text="Refresh",
                            width=7, command=Start)
        refreshBtn.grid(
            row=0, column=8, padx=5, pady=12)

        excel()

        # =======================TABLE FRAME========================

        Table_Frame = Frame(Detail_Frame, bd=4, relief=RIDGE, bg="green4")
        Table_Frame.place(x=20, y=50, width=880, height=640)

        scroll_x = Scrollbar(Table_Frame, orient=HORIZONTAL)
        scroll_y = Scrollbar(Table_Frame, orient=VERTICAL)
        Student_table = ttk.Treeview(Table_Frame, style="mystyle.Treeview", columns=("Roll", "Name", "Class", "Section", "F.Name",
                                                                                     "M.Name", "Contact", "Address", "Branch", "House", "D.O.B", "Aadhar", "Caste", "Category", "BPL"), xscrollcommand=scroll_x.set, yscrollcommand=scroll_y.set)

        scroll_x.pack(side=BOTTOM, fill=X)
        scroll_y.pack(side=RIGHT, fill=Y)
        scroll_x.config(command=Student_table.xview)
        scroll_y.config(command=Student_table.yview)

        Student_table.heading("Roll", text="Roll No.")
        Student_table.heading("Name", text="Name")
        Student_table.heading("Class", text="Class")
        Student_table.heading("Section", text="Section")
        Student_table.heading("F.Name", text="F.Name")
        Student_table.heading("M.Name", text="M.Name")
        Student_table.heading("Contact", text="Contact")
        Student_table.heading("Address", text="Address")
        Student_table.heading("Branch", text="Branch")
        Student_table.heading("House", text="House")
        Student_table.heading("D.O.B", text="D.O.B")
        Student_table.heading("Aadhar", text="Aadhar No.")
        Student_table.heading("Caste", text="Caste")
        Student_table.heading("Category", text="Category")
        Student_table.heading("BPL", text="BPL")

        Student_table['show'] = 'headings'

        Student_table.column("Roll", width=50, anchor="center")
        Student_table.column("Name", width=100, anchor="center")
        Student_table.column("Class", width=100, anchor="center")
        Student_table.column("Section", width=100, anchor="center")
        Student_table.column("F.Name", width=100, anchor="center")
        Student_table.column("M.Name", width=100, anchor="center")
        Student_table.column("Contact", width=100, anchor="center")
        Student_table.column("Address", width=200, anchor="center")
        Student_table.column("Branch", width=100, anchor="center")
        Student_table.column("House", width=100, anchor="center")
        Student_table.column("D.O.B", width=100, anchor="center")
        Student_table.column("Aadhar", width=150, anchor="center")
        Student_table.column("Caste", width=100, anchor="center")
        Student_table.column("Category", width=100, anchor="center")
        Student_table.column("BPL", width=100, anchor="center")

        Student_table.pack(fill=BOTH, expand=1)
        style = ttk.Style()
        style.configure("mystyle.Treeview.Heading",
                        font=('Calibri', 10, 'bold'))

        Student_table.bind("<ButtonRelease-1>", get_cursor)

        Start()

        # ==================STUDENT DETAILS=========================

        def printPdf():
            if(txt_Roll.get() == "" or
                    txt_name.get() == "" or
                    txt_class.get() == "" or
                    txt_sec.get() == "" or
                    txt_fname.get() == "" or
                    txt_mname.get() == "" or
                    txt_contact.get() == "" or
                    txt_address.get() == "" or
                    txt_branch.get() == "" or
                    txt_house.get() == "" or
                    txt_dob.get() == "" or
                    txt_aadhar.get() == "" or
                    txt_caste.get() == "" or
                    txt_category.get() == "" or
                    txt_bpl.get() == ""):

                messagebox.showerror("Error", "Data is not Found for Print")

            else:
                pdf = FPDF()
                # Create the special value {nb}
                pdf.alias_nb_pages()
                pdf.add_page()
                pdf.set_font('Times', '', 12)
                stName = self.name_var.get()
                stRoll = self.roll_no_var.get()
                stClass = self.class_var.get()
                stSec = self.section_var.get()
                stFname = self.fname_var.get()
                stMname = self.mname_var.get()
                stContact = self.contact_var.get()
                stAdd = self.address_var.get()
                stBranch = self.branch_var.get()
                stHouse = self.house_var.get()
                stDob = self.dob_var.get()
                stAdhar = self.aadhar_var.get()
                stCaste = self.caste_var.get()
                stCategory = self.category_var.get()
                stBpl = self.bpl_var.get()

                pdf.cell(0, 10, txt="Student Details", align='C', ln=1)
                pdf.cell(0, 10, txt="{0}: {1}".format(
                    "Name", stName), align='C', ln=2)
                pdf.cell(0, 10, txt="{0}: {1}".format(
                    "Class", stClass), align='C', ln=3)
                pdf.cell(0, 10, txt="{0}: {1}".format(
                    "Roll No.", stRoll), align='C', ln=4)
                pdf.cell(0, 10, txt="{0}: {1}".format(
                    "Section", stSec), align='C', ln=5)
                pdf.cell(0, 10, txt="{0}: {1}".format(
                    "Father's Name", stFname), align='C', ln=6)
                pdf.cell(0, 10, txt="{0}: {1}".format(
                    "Mother's Name", stMname), align='C', ln=7)
                pdf.cell(0, 10, txt="{0}: {1}".format(
                    "Contact", stContact), align='C', ln=8)
                pdf.cell(0, 10, txt="{0}: {1}".format(
                    "Address", stAdd), align='C', ln=9)
                pdf.cell(0, 10, txt="{0}: {1}".format(
                    "Branch", stAdd), align='C', ln=10)
                pdf.cell(0, 10, txt="{0}: {1}".format(
                    "House", stAdd), align='C', ln=11)
                pdf.cell(0, 10, txt="{0}: {1}".format(
                    "D.O.B", stDob), align='C', ln=12)
                pdf.cell(0, 10, txt="{0}: {1}".format(
                    "Aadhar No.", stAdhar), align='C', ln=13)
                pdf.cell(0, 10, txt="{0}: {1}".format(
                    "Caste", stCaste), align='C', ln=14)
                pdf.cell(0, 10, txt="{0}: {1}".format(
                    "Category", stCategory), align='C', ln=15)
                pdf.cell(0, 10, txt="{0}: {1}".format(
                    "BPL", stBpl), align='C', ln=16)

                pdf.output("{0}_cls{1}_roll{2}.pdf".format(
                    stName, stClass, stRoll))

                clear()

        printBtn = Button(btn_frame, text="Print", width=7,
                          bg="purple1", command=printPdf)
        printBtn.grid(
            row=0, column=4, padx=5, pady=7)


root = Tk()
ob = Student(root)
root.mainloop()
